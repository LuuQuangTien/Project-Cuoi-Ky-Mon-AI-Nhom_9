import pygame
from Button import Button
from Pathfinding import BFS, DFS, UCS, Greedy, SIMULATED_ANEALING, BEAM, SENSORLESS, PARTIALLY_OBSERVABLE, BACKTRACKING, AC_3
from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment

#__________________________Game definition___________________________

#Clolor
BLACK = (0, 0, 0)
WHITE = (255, 255, 255)
GREEN = (173, 204, 96)
RED = (253, 145, 145)
LIGHT_BLUE = (173, 216, 230)
LIGHT_GRAY = (211,211,211)
HIGHDANGER_RED = (197, 21, 21, 60)
MEDDANGER_ORANGE = (209, 110, 1, 60)
LOWDANGER_GREEN = (67, 143, 78, 60)

#Button
button_x = 475
button_y = 325
button_width = 250
button_height = 100
algorithm_button_width = 200
algorithm_button_height = 75
button_gap = 20

ALGORITHMS = {
    "BFS": "BFS",
    "DFS": "DFS",
    "UCS": "UCS",
    "GREEDY": "GREEDY",
    "BEAM": "BEAM",
    "SIM A": "SIMULATED_ANEALING",
    "SEN_LESS": "SENSORLESS",
    "PAR_OBS": "PARTIALLY_OBSERVABLE",
    "BACK": "BACKTRACKING",
    "AC_3": "AC_3",
}

METHOD_LAYOUT = [
    ["BFS", "DFS"],
    ["UCS", "GREEDY"],
    ["BEAM", "SIM A"],
    ["SEN_LESS", "PAR_OBS"],
    ["BACK", "AC_3"]
]

ALGORITHMS_SHEETS = {
    "BFS": "BFS",
    "DFS": "DFS",
    "UCS": "UCS",
    "GREEDY": "Greedy",
    "BEAM": "Beam",
    "SIMULATED_ANEALING": "Simulated_Annealing",
    "SENSORLESS": "Sensorless",
    "PARTIALLY_OBSERVABLE": "Partially_Observable",
    "BACKTRACKING": "Backtracking",
    "AC_3": "AC_3",
}

file = "Stats.xlsx"

#_____________________________Game Class_____________________________
class Menu():
    def __init__(self, screen, tile_size, sprite_sheet):
        self.screen = screen
        self.screen_with, self.screen_height = self.screen.get_size()
        self.tile_size = tile_size
        self.screen_rows, self.screen_cols = self.screen_height // tile_size, self.screen_with // tile_size
        self.font = pygame.font.Font(None, 50)
        self.algorithm_font = pygame.font.Font(None, 40)
        self.sprite_sheet = sprite_sheet
        self.grass_sprite = self.get_sprite(3, 0, "img/Grass_img.png")
        self.stone_sprite = self.get_sprite(0, 1, "img/Stone_img.png")
        self.method = None
        self.method_sheet = None
        self.algorithm_buttons = {}
        self.mode = None
        self.current_solver = None
        self.current_solver_method = None

        #___MENU___
        self.rect_play = Button(self.screen, self.font, button_x, button_y - 75, button_width, button_height, "PLAY", GREEN, BLACK)
        self.rect_computer = Button(self.screen, self.font, button_x, button_y + 75, button_width, button_height, "COMPUTER", GREEN, BLACK)
        #___GAMEOVER___
        self.rect_startmenu = Button(self.screen, self.font, button_x - (50), button_y + 75, button_width + 100, button_height, "RETURN", WHITE, BLACK)
        self.rect_replay = Button(self.screen, self.font, button_x, button_y - 75, button_width, button_height, "REPLAY", WHITE, BLACK)
        #___COMPUTER
        self.rect_return = Button(self.screen, self.font, (self.screen_with - (50 + button_width)),
                                  (self.screen_height - (50 + button_height)), button_width, button_height, "RETURN", GREEN, BLACK)
        self.rect_startplaying = Button(self.screen, self.font, 50, self.screen_height - (50 + button_height), button_width, button_height, "START", RED, BLACK)
        self.rect_static_apple = Button(self.screen, self.font, (self.screen_with - (50 + button_width)*2), (self.screen_height - (50 + button_height)), button_width, button_height, "Test", GREEN, BLACK)

        for col_i, col in enumerate(METHOD_LAYOUT):
            x = 20 + col_i * (algorithm_button_width + button_gap + 20)
            for row_i, button_text in enumerate(col):
                y = 100 + row_i * (algorithm_button_height + button_gap)
                method_name = ALGORITHMS[button_text]
                button = Button(
                    self.screen, self.algorithm_font, x, y,
                    algorithm_button_width, algorithm_button_height,
                    button_text, GREEN, BLACK
                )
                self.algorithm_buttons[method_name] = button


        self.rect_visualize_path = Button(self.screen, self.font, button_x - 50, button_y - 50, algorithm_button_width + 100, algorithm_button_height, "VISUALIZE PATH", GREEN, BLACK)

    def draw_game_grid(self, tile_size, snake):
        self.screen.fill(GREEN)

        for row in range(0, self.screen_height, self.tile_size):
            for col in range(0, self.screen_with, self.tile_size):
                self.screen.blit(self.grass_sprite, (col, row))

        for index, tile in enumerate(snake.wall_pos):
            x = int(tile.x * self.tile_size)
            y = int(tile.y * self.tile_size)
            self.screen.blit(self.stone_sprite, (x, y))

    def get_sprite(self, x, y, img):
        rect = pygame.Rect(x * 48, y * 48, 48, 48)
        grass_sprite_sheet = pygame.image.load(img).convert_alpha()
        sprite = grass_sprite_sheet.subsurface(rect).copy()
        sprite = pygame.transform.smoothscale(sprite, (self.tile_size, self.tile_size))
        return sprite

    def draw_start(self):
        pygame.display.set_caption("Start Menu")
        background = pygame.Surface((self.screen_with, self.screen_height))
        background.fill(GREEN)
        background.fill(LIGHT_BLUE, pygame.Rect(self.screen_with // 3, 0, self.screen_with // 3, 750))
        self.screen.blit(background, (0, 0))

    def draw_gameover(self, gameover_screen):
        pygame.display.set_caption("Game Over")
        background = pygame.Surface((self.screen_with, self.screen_height), pygame.SRCALPHA)
        background.fill(LIGHT_GRAY)
        background.set_alpha(128)
        self.screen.blit(gameover_screen, (0, 0))
        self.screen.blit(background, (0, 0))

    def draw_computer(self):
        pygame.display.set_caption("Computer pathfinding algorithm mode")
        background = pygame.Surface((self.screen_with, self.screen_height))
        background.fill(LIGHT_BLUE, pygame.Rect(0, 0, self.screen_with, self.screen_height // 2))
        background.fill(LIGHT_GRAY, pygame.Rect(0, self.screen_height // 2, self.screen_with, self.screen_height // 2))
        self.screen.blit(background, (0, 0))

    def reset_sheet(self, file, sheet):
        wb = load_workbook(file)
        current_sheet = None
        sheet_index = None
        header_row = 2
        headers = ["Apple", "Pos", "Eaten", "States", "Solution Path"]

        if(sheet in wb.sheetnames):
            current_sheet = wb[sheet]
            sheet_index = wb.worksheets.index(current_sheet)
            wb.remove(current_sheet)
        if(sheet_index is not None): sh = wb.create_sheet(sheet, index=sheet_index)
        else: sh = wb.create_sheet(sheet)
        for header_col, header in enumerate(headers, start=1):
            sh.cell(row=header_row, column=header_col).value = header

        wb.save(file)

    def log_stats(self, file, sheet, apple, eaten, state_num, solution):
        wb = load_workbook(file)
        sh = wb[sheet]
        data_row = 3
        summary_row = len(apple.static_pos) + 5

        while sh.cell(row=data_row, column=1).value is not None: data_row += 1
        apple_idx = apple.static_pos_index
        if(apple_idx < 0): apple_idx = 0
        apple_pos_vector2 = apple.static_pos[apple_idx]
        apple_pos = f"({int(apple_pos_vector2.x), int(apple_pos_vector2.y)})"

        test_stats = [apple_idx + 1, apple_pos, eaten, state_num, solution]
        for col, val in enumerate(test_stats, start=1):
            C = sh.cell(row = data_row, column = col, value = val)
            if(col == 5): C.alignment = Alignment(vertical='center', wrapText=True)
            else: C.alignment = Alignment(horizontal='center', vertical='center')
        sh.column_dimensions['E'].width = 50

        data_end_row = data_row
        range_to_eaten = f'C3:C{data_end_row}'
        range_to_average = f'D3:D{data_end_row}'
        average = f'=SUM({range_to_average}) / COUNTIF({range_to_eaten}, True)'
        sh.cell(row=summary_row, column=1, value="Average explored states:")
        avg_cell = sh.cell(row=summary_row, column=4, value=average)
        avg_cell.number_format = numbers.FORMAT_NUMBER_00

        wb.save(file)

    def use_method(self, method, select):
        if(method == select): return None
        else: return select

    def select_method(self, snake, apple):
        solution = None
        match(self.method):
            case("BFS"):
                bfs = BFS(snake, apple.position)
                solution = bfs.Solving()
                if(solution):
                    if(self.mode == "STATIC_APPLE"):
                        solution_str = " -> ".join([f"({int(v.x)}, {int(v.y)})" for v in solution])
                        self.log_stats(file, "BFS", apple, True, bfs.state_num, solution_str)
                else:
                    bfs_tail = BFS(snake, snake.snake[-1])
                    solution = bfs_tail.Solving()
            case("DFS"):
                dfs = DFS(snake, apple.position)
                solution = dfs.Solving()
                if(solution):
                    if (self.mode == "STATIC_APPLE"):
                        solution_str = " -> ".join([f"({int(v.x)}, {int(v.y)})" for v in solution])
                        self.log_stats(file, "DFS", apple, True, dfs.state_num, solution_str)
                else:
                    dfs_tail = DFS(snake, snake.snake[-1])
                    solution = dfs_tail.Solving()
            case("UCS"):
                ucs = UCS(self.screen, self.tile_size, snake, apple.position)
                solution = ucs.Solving()
                if(solution):
                    if (self.mode == "STATIC_APPLE"):
                        solution_str = " -> ".join([f"({int(v.x)}, {int(v.y)})" for v in solution])
                        self.log_stats(file, "UCS", apple, True, ucs.state_num, solution_str)
                else:
                    ucs_tail = UCS(self.screen, self.tile_size, snake, snake.snake[-1])
                    solution = ucs_tail.Solving()
            case("GREEDY"):
                greedy = Greedy(snake, apple.position)
                solution = greedy.Solving()
                if (solution):
                    if (self.mode == "STATIC_APPLE"):
                        solution_str = " -> ".join([f"({int(v.x)}, {int(v.y)})" for v in solution])
                        self.log_stats(file, "Greedy", apple, True, greedy.state_num, solution_str)
                else:
                    greedy_tail = Greedy(snake, snake.snake[-1])
                    solution = greedy_tail.Solving()
            case("SIMULATED_ANEALING"):
                sa = SIMULATED_ANEALING(snake, apple.position, T = 1000, alpha = 0.9)
                solution = sa.Solving()
                if (solution):
                    apple_eaten = (solution[-1].x == apple.position.x and
                                   solution[-1].y == apple.position.y)
                    if (self.mode == "STATIC_APPLE"):
                        solution_str = " -> ".join([f"({int(v.x)}, {int(v.y)})" for v in solution])
                        self.log_stats(file, "Simulated_Annealing", apple, apple_eaten, sa.state_num, solution_str)
                else:
                    sa_tail = SIMULATED_ANEALING(snake, snake.snake[-1], T = 1000, alpha = 0.9)
                    solution = sa_tail.Solving()
            case("BEAM"):
                beam = BEAM(snake, apple.position, 3)
                solution = beam.Solving()
                if (solution):
                    if (self.mode == "STATIC_APPLE"):
                        solution_str = " -> ".join([f"({int(v.x)}, {int(v.y)})" for v in solution])
                        self.log_stats(file, "Beam", apple, True, beam.state_num, solution_str)
                else:
                    beam_tail = BEAM(snake, snake.snake[-1], 5)
                    solution = beam_tail.Solving()
            case("SENSORLESS"):
                sensorless = SENSORLESS(snake, apple.position)
                solution = sensorless.Solving()
                if not(solution):
                    sensorless_tail = SENSORLESS(snake, snake.snake[-1])
                    solution = sensorless_tail.Solving()
            case("PARTIALLY_OBSERVABLE"):
                pa_ob = PARTIALLY_OBSERVABLE(snake, apple.position, grid_size = self.screen_rows * self.screen_cols)
                solution = pa_ob.Solving()
                if (solution):
                    if (self.mode == "STATIC_APPLE"):
                        solution_str = " -> ".join([f"({int(v.x)}, {int(v.y)})" for v in solution])
                        self.log_stats(file, "Partially_Observable", apple, True, pa_ob.state_num, solution_str)
                else:
                    pa_ob_tail = PARTIALLY_OBSERVABLE(snake, snake.snake[-1], grid_size = self.screen_rows * self.screen_cols)
                    solution = pa_ob_tail.Solving()
            case ("BACKTRACKING"):
                back = BACKTRACKING(snake, apple.position)
                solution = back.Solving()
                if (solution):
                    if (self.mode == "STATIC_APPLE"):
                        solution_str = " -> ".join([f"({int(v.x)}, {int(v.y)})" for v in solution])
                        self.log_stats(file, "Backtracking", apple, True, back.nodes_visited, solution_str)
                else:
                    back_tail = BACKTRACKING(snake, snake.snake[-1])
                    solution = back_tail.Solving()
            case("AC_3"):
                ac = AC_3(snake, apple.position)
                solution = ac.Solving()
                if not (solution):
                    ac_tail = AC_3(snake, snake.snake[-1])
                    solution = ac_tail.Solving()

        if(solution and len(solution) > 1):
            solution.popleft()
            return solution
        return None

    def update_start_button_color(self):
        new_color = GREEN if self.method is not None else RED
        self.rect_startplaying.button_color = new_color

    def highlight_dangerzone(self, snake):
        if(self.method == "UCS"):
            for x in range(self.screen_cols):
                for y in range(self.screen_rows):
                    pos = (x, y)
                    if(pos in snake.high_dangerzone): color = HIGHDANGER_RED
                    elif(pos in snake.med_dangerzone): color = MEDDANGER_ORANGE
                    elif(pos in snake.low_dangerzone): color = LOWDANGER_GREEN
                    elif(pos in snake.verylow_dangerzone): color = (173, 204, 96, 0)

                    rect = pygame.Rect(x * self.tile_size, y * self.tile_size, self.tile_size, self.tile_size)
                    zone = pygame.Surface((self.tile_size, self.tile_size), pygame.SRCALPHA)
                    zone.fill(color)
                    self.screen.blit(zone, rect.topleft)

    def input(self, state, events, apple):
        for event in events:
            if(event.type == pygame.QUIT):
                return "QUIT"

        if(state == "MENU"):
            if(self.rect_play.draw(events, "normal")):
                print("clicked PLAY")
                state = "PLAY"
            elif(self.rect_computer.draw(events, "normal")):
                print("clicked COMPUTER")
                state = "COMPUTER"
        elif(state == "GAMEOVER"):
            if(self.rect_startmenu.draw(events, "normal")):
                print("clicked RETURN MENU")
                state = "MENU"
            elif(self.rect_replay.draw(events, "normal")):
                print("clicked REPLAY")
                if(self.method is None): state = "PLAY"
                else: state = "SIMULATE"
        elif(state == "COMPUTER"):
            for method_name, button in self.algorithm_buttons.items():
                pressed = button.draw(events, "press")
                if pressed:
                    # reset tất cả nút khác, không cho các nút đậm cùng lúc
                    for other_name, other_button in self.algorithm_buttons.items():
                        if other_name != method_name:
                            other_button.clicked = False
                    button.clicked = True
                    self.method = method_name
                    self.method_sheet = ALGORITHMS_SHEETS[self.method]
            if (self.rect_startplaying.draw(events, "normal")):
                if(self.method is not None):
                    print("clicked STARTPLAY")
                    if(self.mode == "STATIC_APPLE"): self.reset_sheet(file, self.method_sheet)
                    state = "SIMULATE"
            elif (self.rect_return.draw(events, "normal")):
                print("clicked RETURN")
                state = "MENU"
            elif(self.rect_static_apple.draw(events, "press")):
                if(self.method is not None): self.mode = "STATIC_APPLE"

                apple.reset_position("STATIC_APPLE")
        return state
        
